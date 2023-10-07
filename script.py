import time
import pygetwindow as gw
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill  # Import the PatternFill class
import signal
import sys
from datetime import datetime

# Initialize variables to store the active window and start time
active_window = None
start_time = None

# Dictionary to store window titles and their respective total time
window_time_dict = {}

# Get today's date
today_date = datetime.now().strftime("%Y-%m-%d")

# Define a signal handler to save the data when the script is interrupted
def save_data_on_exit(signal, frame):
    global window_time_dict
    df = pd.DataFrame.from_dict(window_time_dict, orient='index', columns=['Total Time (s)'])
    
    # Calculate the total time in hours
    total_hours = df['Total Time (s)'].sum() / 3600  # Convert seconds to hours
    
    # Add a "Total Hour" row at the bottom with the same number of columns
    df.loc['Total Hour'] = [total_hours]
    
    # Append the date to the "Total Hour" row
    df.at['Total Hour', 'Date'] = today_date
    
    current_directory = os.getcwd()
    file_path = os.path.join(current_directory, 'window_time_data.xlsx')
    
    if os.path.exists(file_path):
        book = load_workbook(file_path)
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            writer.book = book
            
            if 'Time Data' in book.sheetnames:
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                df.to_excel(writer, 'Time Data', index=True, header=True)
                
                # Get the worksheet and set the background color for the "Total Hour" row
                ws = writer.sheets['Time Data']
                total_hour_row = ws.max_row  # Assuming "Total Hour" is the last row
                for cell in ws[f"A{total_hour_row}:Z{total_hour_row}"]:
                    for col_cell in cell:
                        col_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            else:
                df.to_excel(writer, 'Time Data', index=True, header=True)
    else:
        df.to_excel(file_path, index=True, header=True)
    
    sys.exit(0)

# Register the signal handler
signal.signal(signal.SIGINT, save_data_on_exit)

# Flag to keep track of whether today's date has been added to the DataFrame
date_added = False

while True:
    new_active_window = gw.getActiveWindow()
    
    if new_active_window != active_window:
        # If the active window has changed
        if start_time is not None:
            # Stop timing the previous active window
            end_time = time.time()
            elapsed_time = end_time - start_time
            
            if active_window is not None:
                if active_window.title in window_time_dict:
                    window_time_dict[active_window.title] += elapsed_time
                else:
                    window_time_dict[active_window.title] = elapsed_time
                
                print(f"Time spent in '{active_window.title}': {elapsed_time:.2f} seconds")
        
        # Start timing the new active window
        active_window = new_active_window
        start_time = time.time()
    
    # Check if today's date has been added to the DataFrame
    if not date_added:
        date_added = True
        print(f"Adding today's date '{today_date}' to the DataFrame.")
    
    # Check every 1 second (adjust this as needed)
    time.sleep(1)
